"""
Excel Import/Export Handler for Fleet Management System
Handles template generation and data import from Excel files
"""

import io
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def create_fleet_template():
    """
    Creates an Excel template with two sheets:
    1. Vehicles - All vehicle information fields
    2. Maintenance Records - All maintenance record fields
    
    Returns: BytesIO object containing the Excel file
    """
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='012638', end_color='012638', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Optional field styling (lighter header)
    optional_fill = PatternFill(start_color='76777B', end_color='76777B', fill_type='solid')
    
    # Example row styling
    example_fill = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
    example_font = Font(italic=True, color='666666')
    
    # ===== VEHICLES SHEET =====
    ws_vehicles = wb.active
    ws_vehicles.title = "Vehicles"
    
    # Vehicle columns with descriptions
    vehicle_columns = [
        ('VIN*', 'Vehicle Identification Number (17 characters)', 20),
        ('Make*', 'Vehicle manufacturer (e.g., Ford, Toyota)', 15),
        ('Model*', 'Vehicle model (e.g., Transit 250)', 20),
        ('Year*', 'Model year (e.g., 2024)', 10),
        ('License Plate*', 'License plate number', 15),
        ('Purchase Date*', 'Date purchased (YYYY-MM-DD)', 15),
        ('Current Mileage', 'Current odometer reading', 15),
        ('Status', 'Active, In Maintenance, or Retired', 15),
        ('Assigned Driver', 'Driver name', 20),
    ]
    
    # Write vehicle headers
    for col, (header, description, width) in enumerate(vehicle_columns, 1):
        cell = ws_vehicles.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill if '*' in header else optional_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_vehicles.column_dimensions[get_column_letter(col)].width = width
        
        # Add description row
        desc_cell = ws_vehicles.cell(row=2, column=col, value=description)
        desc_cell.font = Font(italic=True, size=9, color='666666')
        desc_cell.alignment = Alignment(wrap_text=True)
    
    # Add example row
    example_vehicle = [
        '1FTYR1ZM5HKB10739',  # VIN
        'Ford',               # Make
        'Transit 250',        # Model
        2024,                 # Year
        'ABC1234',            # License Plate
        '2024-01-15',         # Purchase Date
        15000,                # Current Mileage
        'Active',             # Status
        'John Smith',         # Assigned Driver
    ]
    
    for col, value in enumerate(example_vehicle, 1):
        cell = ws_vehicles.cell(row=3, column=col, value=value)
        cell.fill = example_fill
        cell.font = example_font
        cell.border = thin_border
    
    # Add status dropdown validation
    status_validation = DataValidation(
        type='list',
        formula1='"Active,In Maintenance,Retired"',
        allow_blank=True
    )
    status_validation.error = 'Please select a valid status'
    status_validation.errorTitle = 'Invalid Status'
    ws_vehicles.add_data_validation(status_validation)
    status_validation.add('H3:H1000')  # Status column
    
    # Freeze header rows
    ws_vehicles.freeze_panes = 'A3'
    
    # ===== MAINTENANCE RECORDS SHEET =====
    ws_maintenance = wb.create_sheet("Maintenance Records")
    
    # Maintenance columns with descriptions
    maintenance_columns = [
        ('Vehicle VIN*', 'VIN of the vehicle (must match Vehicles sheet)', 20),
        ('Maintenance Type*', 'Type of service performed', 18),
        ('Service Date*', 'Date of service (YYYY-MM-DD)', 15),
        ('Mileage at Service*', 'Odometer reading at service', 18),
        ('Cost', 'Total cost of service ($)', 12),
        ('Service Provider', 'Name of service provider/shop', 25),
        ('Notes', 'Additional notes or details', 40),
        ('Next Service Due', 'Date of next service (YYYY-MM-DD)', 18),
        ('Next Service Mileage', 'Mileage for next service', 18),
    ]
    
    # Write maintenance headers
    for col, (header, description, width) in enumerate(maintenance_columns, 1):
        cell = ws_maintenance.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill if '*' in header else optional_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws_maintenance.column_dimensions[get_column_letter(col)].width = width
        
        # Add description row
        desc_cell = ws_maintenance.cell(row=2, column=col, value=description)
        desc_cell.font = Font(italic=True, size=9, color='666666')
        desc_cell.alignment = Alignment(wrap_text=True)
    
    # Add example rows
    example_maintenance = [
        ['1FTYR1ZM5HKB10739', 'Oil Change', '2024-01-20', 15500, 45.99, 'Quick Lube', 'Full synthetic oil', '2024-04-20', 18500],
        ['1FTYR1ZM5HKB10739', 'Tire Rotation', '2024-01-20', 15500, 25.00, 'Quick Lube', 'Rotated all 4 tires', '', ''],
        ['1FTYR1ZM5HKB10739', 'Brake Inspection', '2024-02-15', 16200, 0.00, 'Dealer Service', 'Brakes at 60% life', '2024-08-15', ''],
    ]
    
    for row_idx, row_data in enumerate(example_maintenance, 3):
        for col, value in enumerate(row_data, 1):
            cell = ws_maintenance.cell(row=row_idx, column=col, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = thin_border
    
    # Add maintenance type dropdown validation
    maintenance_types = '"Oil Change,Tire Rotation,Tire Replacement,Brake Service,Brake Inspection,Inspection,Repair,Scheduled Maintenance,Battery,Transmission,Engine,Smog Check,Registration,Other"'
    type_validation = DataValidation(
        type='list',
        formula1=maintenance_types,
        allow_blank=False
    )
    type_validation.error = 'Please select a maintenance type'
    type_validation.errorTitle = 'Invalid Type'
    ws_maintenance.add_data_validation(type_validation)
    type_validation.add('B3:B1000')  # Maintenance Type column
    
    # Freeze header rows
    ws_maintenance.freeze_panes = 'A3'
    
    # ===== INSTRUCTIONS SHEET =====
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions = [
        ("Fleet Data Import Template", None),
        ("", None),
        ("INSTRUCTIONS:", None),
        ("1. Fill in your vehicle data in the 'Vehicles' sheet", None),
        ("2. Fill in maintenance records in the 'Maintenance Records' sheet", None),
        ("3. Fields marked with * are required", None),
        ("4. Delete the example rows (highlighted in green) before importing", None),
        ("5. Save the file and upload it to the Fleet Manager", None),
        ("", None),
        ("IMPORTANT NOTES:", None),
        ("• VINs must be unique and exactly 17 characters", None),
        ("• License plates must be unique", None),
        ("• Dates must be in YYYY-MM-DD format (e.g., 2024-01-15)", None),
        ("• Maintenance records must reference a valid VIN from the Vehicles sheet", None),
        ("• Cost values should be numbers only (no $ symbol)", None),
        ("", None),
        ("VEHICLE STATUS OPTIONS:", None),
        ("• Active - Vehicle is in regular use", None),
        ("• In Maintenance - Vehicle is currently being serviced", None),
        ("• Retired - Vehicle is no longer in service", None),
        ("", None),
        ("MAINTENANCE TYPES:", None),
        ("• Oil Change", None),
        ("• Tire Rotation", None),
        ("• Tire Replacement", None),
        ("• Brake Service", None),
        ("• Brake Inspection", None),
        ("• Inspection", None),
        ("• Repair", None),
        ("• Scheduled Maintenance", None),
        ("• Battery", None),
        ("• Transmission", None),
        ("• Engine", None),
        ("• Smog Check", None),
        ("• Registration", None),
        ("• Other", None),
    ]
    
    # Write instructions
    title_font = Font(bold=True, size=16, color='012638')
    section_font = Font(bold=True, size=12, color='012638')
    
    for row_idx, (text, _) in enumerate(instructions, 1):
        cell = ws_instructions.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = title_font
        elif text.endswith(':'):
            cell.font = section_font
    
    ws_instructions.column_dimensions['A'].width = 70
    
    # Move Instructions to first position
    wb.move_sheet(ws_instructions, offset=-2)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def parse_excel_import(file_stream):
    """
    Parses an uploaded Excel file and extracts vehicle and maintenance data.
    
    Args:
        file_stream: File object or BytesIO containing Excel data
        
    Returns:
        dict with 'vehicles', 'maintenance', 'errors', and 'warnings'
    """
    result = {
        'vehicles': [],
        'maintenance': [],
        'errors': [],
        'warnings': []
    }
    
    try:
        wb = load_workbook(file_stream)
    except Exception as e:
        result['errors'].append(f"Could not read Excel file: {str(e)}")
        return result
    
    # ===== PARSE VEHICLES =====
    if 'Vehicles' in wb.sheetnames:
        ws = wb['Vehicles']
        
        # Find the data starting row (skip header and description rows)
        for row_idx in range(3, ws.max_row + 1):
            # Skip empty rows
            vin = ws.cell(row=row_idx, column=1).value
            if not vin or str(vin).strip() == '':
                continue
            
            # Skip example row (check if it's the example VIN)
            if str(vin).strip() == '1FTYR1ZM5HKB10739':
                # Check if it's the example by looking at other fields
                make = ws.cell(row=row_idx, column=2).value
                model = ws.cell(row=row_idx, column=3).value
                if make == 'Ford' and model == 'Transit 250':
                    result['warnings'].append(f"Row {row_idx}: Skipped example row")
                    continue
            
            vehicle = {
                'vin': str(vin).strip().upper(),
                'make': str(ws.cell(row=row_idx, column=2).value or '').strip(),
                'model': str(ws.cell(row=row_idx, column=3).value or '').strip(),
                'year': ws.cell(row=row_idx, column=4).value,
                'license_plate': str(ws.cell(row=row_idx, column=5).value or '').strip().upper(),
                'purchase_date': ws.cell(row=row_idx, column=6).value,
                'current_mileage': ws.cell(row=row_idx, column=7).value or 0,
                'status': str(ws.cell(row=row_idx, column=8).value or 'Active').strip(),
                'assigned_driver': str(ws.cell(row=row_idx, column=9).value or '').strip(),
                'row': row_idx
            }
            
            # Validate required fields
            errors = []
            if not vehicle['vin']:
                errors.append("VIN is required")
            elif len(vehicle['vin']) != 17:
                errors.append(f"VIN must be 17 characters (got {len(vehicle['vin'])})")
            
            if not vehicle['make']:
                errors.append("Make is required")
            if not vehicle['model']:
                errors.append("Model is required")
            if not vehicle['year']:
                errors.append("Year is required")
            elif not isinstance(vehicle['year'], int):
                try:
                    vehicle['year'] = int(vehicle['year'])
                except:
                    errors.append("Year must be a number")
            
            if not vehicle['license_plate']:
                errors.append("License plate is required")
            
            if not vehicle['purchase_date']:
                errors.append("Purchase date is required")
            else:
                # Parse date
                if isinstance(vehicle['purchase_date'], datetime):
                    vehicle['purchase_date'] = vehicle['purchase_date'].date()
                elif isinstance(vehicle['purchase_date'], str):
                    try:
                        vehicle['purchase_date'] = datetime.strptime(vehicle['purchase_date'], '%Y-%m-%d').date()
                    except:
                        errors.append("Purchase date must be in YYYY-MM-DD format")
            
            # Parse mileage
            if vehicle['current_mileage']:
                try:
                    vehicle['current_mileage'] = int(vehicle['current_mileage'])
                except:
                    errors.append("Mileage must be a number")
            else:
                vehicle['current_mileage'] = 0
            
            # Validate status
            valid_statuses = ['Active', 'In Maintenance', 'Retired']
            if vehicle['status'] not in valid_statuses:
                errors.append(f"Status must be one of: {', '.join(valid_statuses)}")
            
            if errors:
                result['errors'].append(f"Row {row_idx} (VIN: {vehicle['vin']}): {'; '.join(errors)}")
            else:
                result['vehicles'].append(vehicle)
    else:
        result['errors'].append("'Vehicles' sheet not found in Excel file")
    
    # ===== PARSE MAINTENANCE RECORDS =====
    if 'Maintenance Records' in wb.sheetnames:
        ws = wb['Maintenance Records']
        
        # Get list of valid VINs from parsed vehicles
        valid_vins = {v['vin'] for v in result['vehicles']}
        
        for row_idx in range(3, ws.max_row + 1):
            # Skip empty rows
            vin = ws.cell(row=row_idx, column=1).value
            if not vin or str(vin).strip() == '':
                continue
            
            # Skip example rows
            if str(vin).strip().upper() == '1FTYR1ZM5HKB10739':
                maint_type = ws.cell(row=row_idx, column=2).value
                if maint_type in ['Oil Change', 'Tire Rotation', 'Brake Inspection']:
                    service_date = ws.cell(row=row_idx, column=3).value
                    if service_date and ('2024-01' in str(service_date) or '2024-02' in str(service_date)):
                        result['warnings'].append(f"Maintenance Row {row_idx}: Skipped example row")
                        continue
            
            record = {
                'vehicle_vin': str(vin).strip().upper(),
                'maintenance_type': str(ws.cell(row=row_idx, column=2).value or '').strip(),
                'service_date': ws.cell(row=row_idx, column=3).value,
                'mileage_at_service': ws.cell(row=row_idx, column=4).value,
                'cost': ws.cell(row=row_idx, column=5).value or 0,
                'service_provider': str(ws.cell(row=row_idx, column=6).value or '').strip(),
                'notes': str(ws.cell(row=row_idx, column=7).value or '').strip(),
                'next_service_due': ws.cell(row=row_idx, column=8).value,
                'next_service_mileage': ws.cell(row=row_idx, column=9).value,
                'row': row_idx
            }
            
            # Validate required fields
            errors = []
            
            if not record['vehicle_vin']:
                errors.append("Vehicle VIN is required")
            elif record['vehicle_vin'] not in valid_vins:
                errors.append(f"VIN '{record['vehicle_vin']}' not found in Vehicles sheet")
            
            if not record['maintenance_type']:
                errors.append("Maintenance type is required")
            
            if not record['service_date']:
                errors.append("Service date is required")
            else:
                if isinstance(record['service_date'], datetime):
                    record['service_date'] = record['service_date'].date()
                elif isinstance(record['service_date'], str):
                    try:
                        record['service_date'] = datetime.strptime(record['service_date'], '%Y-%m-%d').date()
                    except:
                        errors.append("Service date must be in YYYY-MM-DD format")
            
            if not record['mileage_at_service']:
                errors.append("Mileage at service is required")
            else:
                try:
                    record['mileage_at_service'] = int(record['mileage_at_service'])
                except:
                    errors.append("Mileage must be a number")
            
            # Parse optional fields
            if record['cost']:
                try:
                    record['cost'] = float(str(record['cost']).replace('$', '').replace(',', ''))
                except:
                    errors.append("Cost must be a number")
            else:
                record['cost'] = 0.0
            
            if record['next_service_due']:
                if isinstance(record['next_service_due'], datetime):
                    record['next_service_due'] = record['next_service_due'].date()
                elif isinstance(record['next_service_due'], str):
                    try:
                        record['next_service_due'] = datetime.strptime(record['next_service_due'], '%Y-%m-%d').date()
                    except:
                        record['next_service_due'] = None
                        result['warnings'].append(f"Maintenance Row {row_idx}: Invalid next service date format, skipping")
            else:
                record['next_service_due'] = None
            
            if record['next_service_mileage']:
                try:
                    record['next_service_mileage'] = int(record['next_service_mileage'])
                except:
                    record['next_service_mileage'] = None
            else:
                record['next_service_mileage'] = None
            
            if errors:
                result['errors'].append(f"Maintenance Row {row_idx}: {'; '.join(errors)}")
            else:
                result['maintenance'].append(record)
    else:
        result['warnings'].append("'Maintenance Records' sheet not found - no maintenance records will be imported")
    
    return result


def import_data_to_db(parsed_data, db, Vehicle, MaintenanceRecord, clear_existing=False):
    """
    Imports parsed data into the database.
    
    Args:
        parsed_data: Result from parse_excel_import()
        db: SQLAlchemy database instance
        Vehicle: Vehicle model class
        MaintenanceRecord: MaintenanceRecord model class
        clear_existing: If True, deletes all existing data before import
        
    Returns:
        dict with 'success', 'vehicles_added', 'maintenance_added', 'errors'
    """
    result = {
        'success': True,
        'vehicles_added': 0,
        'vehicles_skipped': 0,
        'maintenance_added': 0,
        'errors': []
    }
    
    try:
        if clear_existing:
            MaintenanceRecord.query.delete()
            Vehicle.query.delete()
            db.session.commit()
        
        # Create a mapping of VIN to vehicle ID for maintenance records
        vin_to_id = {}
        
        # First, check for existing vehicles
        existing_vins = {v.vin for v in Vehicle.query.all()}
        existing_plates = {v.license_plate for v in Vehicle.query.all()}
        
        # Import vehicles
        for v_data in parsed_data['vehicles']:
            # Skip if VIN already exists (unless clearing)
            if v_data['vin'] in existing_vins and not clear_existing:
                result['warnings'] = result.get('warnings', [])
                result['warnings'].append(f"Vehicle with VIN {v_data['vin']} already exists - skipped")
                result['vehicles_skipped'] += 1
                # Still add to mapping for maintenance records
                existing_vehicle = Vehicle.query.filter_by(vin=v_data['vin']).first()
                if existing_vehicle:
                    vin_to_id[v_data['vin']] = existing_vehicle.id
                continue
            
            # Skip if license plate already exists
            if v_data['license_plate'] in existing_plates and not clear_existing:
                result['errors'].append(f"Vehicle with license plate {v_data['license_plate']} already exists - skipped")
                result['vehicles_skipped'] += 1
                continue
            
            vehicle = Vehicle(
                vin=v_data['vin'],
                make=v_data['make'],
                model=v_data['model'],
                year=v_data['year'],
                license_plate=v_data['license_plate'],
                purchase_date=v_data['purchase_date'],
                current_mileage=v_data['current_mileage'],
                status=v_data['status'],
                assigned_driver=v_data['assigned_driver'] if v_data['assigned_driver'] else None
            )
            
            db.session.add(vehicle)
            db.session.flush()  # Get the ID
            vin_to_id[v_data['vin']] = vehicle.id
            existing_vins.add(v_data['vin'])
            existing_plates.add(v_data['license_plate'])
            result['vehicles_added'] += 1
        
        # Also map existing vehicles for maintenance records
        for vehicle in Vehicle.query.all():
            if vehicle.vin not in vin_to_id:
                vin_to_id[vehicle.vin] = vehicle.id
        
        # Import maintenance records
        for m_data in parsed_data['maintenance']:
            vehicle_id = vin_to_id.get(m_data['vehicle_vin'])
            
            if not vehicle_id:
                result['errors'].append(f"Could not find vehicle for VIN {m_data['vehicle_vin']}")
                continue
            
            record = MaintenanceRecord(
                vehicle_id=vehicle_id,
                maintenance_type=m_data['maintenance_type'],
                service_date=m_data['service_date'],
                mileage_at_service=m_data['mileage_at_service'],
                cost=m_data['cost'],
                service_provider=m_data['service_provider'] if m_data['service_provider'] else None,
                notes=m_data['notes'] if m_data['notes'] else None,
                next_service_due=m_data['next_service_due'],
                next_service_mileage=m_data['next_service_mileage']
            )
            
            db.session.add(record)
            result['maintenance_added'] += 1
        
        db.session.commit()
        
    except Exception as e:
        db.session.rollback()
        result['success'] = False
        result['errors'].append(f"Database error: {str(e)}")
    
    return result
